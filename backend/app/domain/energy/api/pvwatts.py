"""PVWatts API endpoints — recalculate, update-input, geocode, and Excel export."""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Any
import io
import logging

from app.core.auth import get_current_user, MockUser
from app.domain.energy.assessments.pvwatts_assessment import PVWattsTool
from app.domain.energy.services.pvwatts_engine import PVWattsEngine, MONTH_LABELS, MODULE_TYPE_LABELS, ARRAY_TYPE_LABELS

router = APIRouter()
logger = logging.getLogger(__name__)


class RecalculateRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]


class UpdateInputRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]
    field_name: str
    value: Any
    source: str = "user"
    status: str = "validated"


def _normalize_input_status(status: str) -> str:
    return status


def _normalize_input_value(value: Any) -> Any:
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    lowered = value.strip().lower()
    if lowered in {"", "—", "-", "–", "n/a", "na", "none", "null", "missing", "unknown"} or lowered.startswith("unknown "):
        return None
    return value.strip()


class GeocodeRequest(BaseModel):
    address: str


@router.post("/pvwatts/recalculate")
async def recalculate_solar(
    data: RecalculateRequest,
    user: MockUser = Depends(get_current_user),
):
    """Recalculate solar estimate from a full input set. Calls PVWatts API."""
    tool = PVWattsTool()
    try:
        result = await tool.recalculate(data.inputs)
        return result
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid inputs for solar calculation")
    except Exception as e:
        logger.error(f"PVWatts recalculate failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Solar calculation failed. Please try again.")


@router.post("/pvwatts/update-input")
async def update_input_and_recalculate(
    data: UpdateInputRequest,
    user: MockUser = Depends(get_current_user),
):
    """Update a single input field and recalculate."""
    inputs = data.inputs
    normalized_value = _normalize_input_value(data.value)
    normalized_status = "missing" if normalized_value is None else _normalize_input_status(data.status)
    if data.field_name in inputs:
        inputs[data.field_name]["value"] = normalized_value
        inputs[data.field_name]["source"] = data.source
        inputs[data.field_name]["status"] = normalized_status
    else:
        inputs[data.field_name] = {
            "field_name": data.field_name,
            "label": data.field_name,
            "value": normalized_value,
            "unit": "",
            "source": data.source,
            "status": normalized_status,
            "notes": "",
            "rationale": "",
            "category": "general",
        }

    tool = PVWattsTool()
    try:
        result = await tool.recalculate(inputs)
        return result
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid inputs for solar calculation")
    except Exception as e:
        logger.error(f"PVWatts update-input failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Solar calculation failed. Please try again.")


@router.post("/pvwatts/geocode")
async def geocode_address(
    data: GeocodeRequest,
    user: MockUser = Depends(get_current_user),
):
    """Geocode an address string to lat/lon coordinates."""
    try:
        result = await PVWattsEngine.geocode_address(data.address)
        return result
    except ValueError:
        raise HTTPException(status_code=404, detail="Location not found")
    except Exception as e:
        logger.error(f"Geocoding failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Geocoding service unavailable. Please try again.")


class AutocompleteRequest(BaseModel):
    query: str


@router.post("/pvwatts/autocomplete")
async def autocomplete_address(
    data: AutocompleteRequest,
    user: MockUser = Depends(get_current_user),
):
    """Autocomplete/typeahead for address search."""
    try:
        results = await PVWattsEngine.autocomplete_address(data.query)
        return {"results": results}
    except Exception as e:
        logger.error(f"Autocomplete failed: {e}", exc_info=True)
        return {"results": []}


class ExportRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]
    result: dict[str, Any]


@router.post("/pvwatts/export")
async def export_solar_excel(
    data: ExportRequest,
    user: MockUser = Depends(get_current_user),
):
    """Export the solar production estimate as an Excel workbook.
    Sheets: Summary (headline metrics + monthly data), Inputs.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    inp = data.inputs
    res = data.result

    def _v(field: str, fallback: Any = None) -> Any:
        return (inp.get(field) or {}).get("value", fallback)

    BLUE = "004D91"
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    title_font = Font(bold=True, size=14, color=BLUE)
    section_font = Font(bold=True, size=11, color=BLUE)
    confirmed_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    assumed_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    int_fmt = "#,##0"
    dec_fmt = "#,##0.0"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = BLUE
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Solar Production Estimate"
    ws["A1"].font = title_font

    address_val = _v("address", "")
    if address_val:
        ws.merge_cells("A2:C2")
        ws["A2"].value = str(address_val)
        ws["A2"].font = Font(italic=True, color="555555", size=10)

    row = 4
    ws.cell(row, 1, "HEADLINE RESULTS").font = section_font
    row += 1

    headline = [
        ("Year 1 AC Energy (kWh)", res.get("ac_annual", 0), int_fmt),
        ("Capacity Factor (%)", res.get("capacity_factor", 0), dec_fmt),
        ("Annual Solar Radiation (kWh/m²/day)", res.get("solrad_annual", 0), dec_fmt),
        ("Data Confidence", res.get("quality_label", ""), None),
    ]
    for label, value, nfmt in headline:
        c_label = ws.cell(row, 1, label)
        c_label.font = Font(size=11)
        c_val = ws.cell(row, 2, round(value, 2) if isinstance(value, float) else value)
        c_val.border = thin
        c_val.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        if nfmt:
            c_val.number_format = nfmt
        row += 1

    row += 1
    ws.cell(row, 1, "MONTHLY PRODUCTION").font = section_font
    row += 1

    for col, hdr in enumerate(["Month", "AC Energy (kWh)", "DC Energy (kWh)", "Solar Radiation (kWh/m²/day)", "POA Irradiance (kWh/m²)"], 1):
        c = ws.cell(row, col, hdr)
        c.font = header_font
        c.fill = header_fill
        c.border = thin
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = max(ws.column_dimensions[c.column_letter].width, len(hdr) + 4)
    row += 1

    ac_monthly = res.get("ac_monthly", [0] * 12)
    dc_monthly = res.get("dc_monthly", [0] * 12)
    solrad_monthly = res.get("solrad_monthly", [0] * 12)
    poa_monthly = res.get("poa_monthly", [0] * 12)

    for i, month in enumerate(MONTH_LABELS):
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") if i % 2 else None
        row_data = [month, round(ac_monthly[i], 0), round(dc_monthly[i], 0), round(solrad_monthly[i], 3), round(poa_monthly[i], 1)]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row, col, val)
            c.border = thin
            if alt_fill:
                c.fill = alt_fill
            if col > 1:
                c.number_format = int_fmt if col <= 3 else dec_fmt
                c.alignment = Alignment(horizontal="right")
        row += 1

    # Annual totals row
    totals = ["Annual", round(sum(ac_monthly), 0), round(sum(dc_monthly), 0),
              round(res.get("solrad_annual", 0), 3), round(sum(poa_monthly), 1)]
    for col, val in enumerate(totals, 1):
        c = ws.cell(row, col, val)
        c.font = Font(bold=True)
        c.border = Border(top=Side(style="double"), left=Side(style="thin"),
                          right=Side(style="thin"), bottom=Side(style="thin"))
        if col > 1:
            c.number_format = int_fmt if col <= 3 else dec_fmt
            c.alignment = Alignment(horizontal="right")

    # Station info footnote
    station = res.get("station_info", {})
    if station.get("state"):
        row += 2
        note = f"Weather data: {station.get('city', '')}, {station['state']}"
        if station.get("weather_data_source"):
            note += f" ({station['weather_data_source']})"
        ws.cell(row, 1, note).font = Font(italic=True, color="888888", size=9)
    row += 1
    ws.cell(row, 1, "Powered by NREL PVWatts V8").font = Font(italic=True, color="888888", size=9)

    # ── Sheet 2: Inputs ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Inputs")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 45

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = "Solar Estimate — Input Parameters"
    ws2["A1"].font = title_font

    for col, hdr in enumerate(["Parameter", "Value", "Unit", "Status", "Notes / Rationale"], 1):
        c = ws2.cell(3, col, hdr)
        c.font = header_font
        c.fill = header_fill
        c.border = thin

    from app.domain.energy.services.pvwatts_engine import CATEGORY_ORDER, CATEGORY_LABELS
    row2 = 4
    for cat in CATEGORY_ORDER:
        cat_rows = [(k, v) for k, v in inp.items() if (v.get("category") or "performance") == cat]
        if not cat_rows:
            continue
        ws2.merge_cells(f"A{row2}:E{row2}")
        c = ws2.cell(row2, 1, CATEGORY_LABELS.get(cat, cat).upper())
        c.font = Font(bold=True, size=10, color=BLUE)
        c.fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        row2 += 1
        for field_name, field in cat_rows:
            val = field.get("value")
            if field_name == "assessment_type":
                val = MODULE_TYPE_LABELS.get(int(val), str(val)) if val is not None else ""
            elif field_name == "array_type":
                val = ARRAY_TYPE_LABELS.get(int(val), str(val)) if val is not None else ""
            status = field.get("status", "")
            row_fill = confirmed_fill if status == "validated" else assumed_fill if status == "assumed" else None
            row_data = [field.get("label", field_name), val, field.get("unit", ""), status, field.get("notes", "")]
            for col, cell_val in enumerate(row_data, 1):
                c = ws2.cell(row2, col, cell_val if cell_val is not None else "—")
                c.border = thin
                if row_fill:
                    c.fill = row_fill
            row2 += 1

    row2 += 1
    ws2.cell(row2, 1, "Green = user-validated  |  Yellow = assumed default").font = Font(italic=True, color="888888", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=solar_estimate.xlsx"},
    )
