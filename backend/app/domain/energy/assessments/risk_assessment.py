"""Risk Assessment Assessment.

Stage workflow:
  1. Categories   (list / categorized_list)
  2. Risks        (list / categorized_workspace)
  3. Mitigations  (record / categorized_workspace)
  4. Register     (computed_results / risk_register_results)

Export: XLSX generated on demand from confirmed register data.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from app.config import get_settings
from app.assessments.base import (
    BaseAssessment,
    DecisionLogAttribution,
    FieldDef,
    AssessmentDefinition,
    AssessmentManifest,
    PopulationStep,
    StageDef,
)
from app.assessments.retrieval import retrieve_evidence
from app.assessments.utils import infer_category_icon, llm_json

logger = logging.getLogger(__name__)
settings = get_settings()

RISK_RATINGS = ("Low", "Moderate", "Substantial", "High")
EVIDENCE_STATUSES = ("Supported", "Partially supported", "Variable", "Needs evidence")

# Coverage targets for the risks stage. Depth is filled by a second LLM pass —
# never by static template text, so output stays project-specific.
MIN_RISKS_PER_CATEGORY = 2
MAX_RISKS_PER_CATEGORY = 4


class RiskAssessment(BaseAssessment):
    """SORT-style risk assessment for sustainable development projects."""

    @property
    def definition(self) -> AssessmentDefinition:
        return AssessmentDefinition(
            id="risk_assessment",
            name="Risk Assessment",
            description="Build a tailored project risk register with mitigations and ratings",
            icon="ShieldAlert",
            output_type="risk_register",
            category="assessment",
            keywords=[
                "risk",
                "risk register",
                "sort",
                "diligence",
                "fiduciary",
                "governance",
                "mitigation",
                "residual risk",
            ],
            export_format="xlsx",
        )

    @property
    def manifest(self) -> AssessmentManifest:
        return AssessmentManifest(
            **self.definition.__dict__,
            goal="Produce a project-specific risk register with mitigations, ratings, evidence basis, and gaps.",
            primary_ui_object="risk_register_results",
            export_artifact_types=["xlsx"],
            adapter_bindings={"research_source": "retrieval"},
            input_dependencies=[],
            produced_outputs=["risk_register"],
            downstream_dependencies=[],
            variables_behavior="tracks",
            evidence_behavior="rag_grounded",
            decision_log_attribution=DecisionLogAttribution(
                adapter_labels={"research_source": "Project materials and research retrieval"},
                widget_detail_labels={
                    "risk_register": "Risk Register",
                    "category_ratings": "Category Ratings",
                    "top_risks": "Top Risks",
                    "unresolved_issues": "Unresolved Issues",
                },
            ),
        )

    @property
    def stage_defs(self) -> list[StageDef]:
        return [
            StageDef(
                id="categories",
                title="Categories",
                component="list",
                widget="categorized_list",
                allow_add_rows=True,
                fields=[
                    FieldDef("label", "text", required=True, label="Category"),
                    FieldDef("description", "long_text", label="Description"),
                    FieldDef("why_it_matters", "long_text", label="Why It Matters"),
                    FieldDef(
                        "status",
                        "select",
                        label="Include?",
                        options=["Include", "Exclude"],
                    ),
                    FieldDef("note", "long_text", label="Note"),
                ],
                population=[
                    PopulationStep("seed_from_template"),
                    PopulationStep("adapt_with_ai_from_project_materials", {"require_citation": True}),
                    PopulationStep("await_user_confirmation"),
                ],
            ),
            StageDef(
                id="risks",
                title="Risks",
                component="list",
                widget="categorized_workspace",
                allow_add_rows=True,
                fields=[
                    FieldDef("title", "text", required=True, label="Risk"),
                    FieldDef("category", "text", required=True, label="Category"),
                    FieldDef("affected_components", "long_text", label="Affected Components"),
                    FieldDef("why_it_matters", "long_text", label="Why It Matters"),
                    FieldDef("evidence_basis", "long_text", label="Evidence / Basis"),
                    FieldDef("missing_information", "long_text", label="Missing Information"),
                    FieldDef(
                        "evidence_status",
                        "select",
                        label="Evidence Status",
                        options=["Supported", "Partially supported", "Variable", "Needs evidence"],
                    ),
                ],
                population=[
                    PopulationStep("read_confirmed_prior_stage", {"stage_id": "categories"}),
                    PopulationStep("extract_from_project_materials"),
                    PopulationStep("propose_with_ai", {"require_citation": True}),
                    PopulationStep("await_user_confirmation"),
                ],
            ),
            StageDef(
                id="mitigations",
                title="Mitigations",
                component="record",
                widget="categorized_workspace",
                fields=[
                    FieldDef("mitigation", "long_text", label="Proposed Mitigation"),
                    FieldDef("owner", "text", label="Likely Owner"),
                    FieldDef("timing", "text", label="Timing / Stage"),
                    FieldDef("remaining_issue", "long_text", label="Remaining Issue"),
                    FieldDef(
                        "status",
                        "select",
                        label="Sufficiency",
                        options=["Adequate", "Insufficient", "Needs validation"],
                    ),
                ],
                population=[
                    PopulationStep("read_confirmed_prior_stage", {"stage_id": "risks"}),
                    PopulationStep("enrich_selected_item_with_ai", {"require_citation": True, "bulk": True}),
                    PopulationStep("await_user_confirmation"),
                ],
            ),
            StageDef(
                id="results",
                title="Results",
                component="computed_results",
                widget="risk_register_results",
                population=[
                    PopulationStep("read_confirmed_prior_stage", {"stage_id": "mitigations"}),
                    PopulationStep("compute_with_assessment_logic"),
                    PopulationStep("await_user_confirmation"),
                ],
            ),
        ]

    async def generate_items_for_stage(
        self,
        stage_id: str,
        step_type: str,
        context: dict,
        prior_data: dict[str, Any],
    ) -> list[dict]:
        if stage_id == "categories":
            return await self._generate_categories(context)
        if stage_id == "risks":
            category_items = (prior_data.get("categories") or {}).get("data", {}).get("items", [])
            return await self._generate_risks(context, category_items)
        return []

    async def enrich_record(
        self,
        stage_id: str,
        item_content: dict,
        existing_record: dict,
        context: dict,
    ) -> dict:
        if stage_id != "mitigations":
            raise ValueError(f"enrich_record called for unexpected stage '{stage_id}'")
        return await self._enrich_mitigation(item_content, existing_record, context)

    async def enrich_records_for_stage(
        self,
        stage_id: str,
        source_items: list[dict[str, Any]],
        existing_records: dict[str, dict[str, Any]],
        context: dict,
    ) -> dict[str, dict[str, Any]]:
        if stage_id != "mitigations":
            raise ValueError(f"enrich_records_for_stage called for unexpected stage '{stage_id}'")
        return await self._enrich_mitigations_bulk(source_items, existing_records, context)

    async def compute_stage(
        self,
        stage_id: str,
        confirmed_stages: dict[str, Any],
        context: dict,
    ) -> dict[str, Any]:
        if stage_id != "results":
            raise ValueError(f"compute_stage called for unexpected stage '{stage_id}'")

        risk_items = (confirmed_stages.get("risks") or {}).get("data", {}).get("items", [])
        category_items = (confirmed_stages.get("categories") or {}).get("data", {}).get("items", [])
        mitigation_records = (confirmed_stages.get("mitigations") or {}).get("data", {}).get("records", {})

        draft_register = self._build_register_seed(risk_items, mitigation_records)
        data = await self._rate_register(context, category_items, draft_register)

        category_ratings = [
            {
                "category": item.get("category", ""),
                "rating": _normalize_rating(item.get("rating")),
                "rationale": item.get("rationale", ""),
                "top_risks": _as_list(item.get("top_risks")),
                "unresolved_issues": _as_list(item.get("unresolved_issues")),
            }
            for item in data.get("category_ratings", [])
            if item.get("category")
        ]

        llm_rows = data.get("risk_register", [])
        register_rows = []
        for idx, seed in enumerate(draft_register):
            llm_row = _find_llm_row(seed, llm_rows, idx)
            inherent = _normalize_rating(llm_row.get("inherent_rating") if llm_row else seed.get("inherent_rating"))
            residual = _normalize_rating(llm_row.get("residual_rating") if llm_row else seed.get("residual_rating"))
            register_rows.append({
                **seed,
                "inherent_rating": inherent,
                "residual_rating": residual,
                "rating_rationale": (llm_row or {}).get("rating_rationale", seed.get("rating_rationale", "")),
                "basis_evidence": (llm_row or {}).get("basis_evidence", seed.get("basis_evidence", "")),
                "missing_information": (llm_row or {}).get("missing_information", seed.get("missing_information", "")),
            })

        if not category_ratings:
            category_ratings = self._derive_category_ratings(category_items, register_rows)

        top_risks = data.get("top_risks") or self._derive_top_risks(register_rows)
        unresolved_issues = data.get("unresolved_issues") or self._derive_unresolved_issues(register_rows)

        return {
            "assessment_id": self.definition.id,
            "rating_scale": list(RISK_RATINGS),
            "project_title": context.get("project_title", ""),
            "category_ratings": category_ratings,
            "risk_register": register_rows,
            "top_risks": top_risks,
            "unresolved_issues": unresolved_issues,
            "copy": {
                "markdown": _risk_register_markdown(register_rows),
                "tsv": _risk_register_tsv(register_rows),
            },
        }

    async def generate_export(self, confirmed_stages: dict[str, Any], context: dict) -> bytes:
        results_data = (confirmed_stages.get("results") or {}).get("data") or {}
        widget_data = results_data.get("widget_data", {})
        return self.export_xlsx(widget_data)

    async def _generate_categories(self, context: dict) -> list[dict]:
        queries = _context_queries(context, "project risk categories SORT development finance")
        evidence_block = await _evidence_block(queries, context)
        data = await llm_json(
            system=(
                "You are a senior development-finance risk reviewer. Generate 6-10 project-specific "
                "risk categories. Use World Bank SORT-style categories only as a starting reference; "
                "adapt categories to the project's geography, sector, delivery model, technology, "
                "beneficiaries, financing, and available evidence. Do not invent facts. Return JSON "
                "with key categories: objects with label, description, why_it_matters, status, note."
            ),
            user_msg=_project_context_text(context) + evidence_block,
            model=settings.openai_orchestration_model,
            context=context,
        )
        categories = data.get("categories") or []
        if not categories:
            categories = _default_categories(context)
        return [
            {
                "label": c.get("label", c.get("name", "")).strip(),
                "description": c.get("description", "").strip(),
                "why_it_matters": c.get("why_it_matters", c.get("rationale", "")).strip(),
                "status": c.get("status") or "Include",
                "note": c.get("note", ""),
                "icon": infer_category_icon(c.get("label", c.get("name", ""))),
            }
            for c in categories
            if c.get("label") or c.get("name")
        ]

    async def _generate_risks(self, context: dict, category_items: list[dict]) -> list[dict]:
        categories = [
            item.get("content", {})
            for item in category_items
            if (item.get("content", {}).get("status") or "Include") != "Exclude"
        ]
        category_text = "\n".join(
            f"- {c.get('label')}: {c.get('why_it_matters') or c.get('description')}"
            for c in categories
        )
        category_labels = [
            str(c.get("label", "")).strip()
            for c in categories
            if str(c.get("label", "")).strip()
        ]
        descriptor = _project_descriptor(context)
        joined_labels = ", ".join(category_labels)
        queries = [
            f"{descriptor} project preparation risks {joined_labels}".strip(),
            f"{descriptor} implementation risks approvals delivery model beneficiaries".strip(),
            f"{descriptor} precedent project risks lessons learned evidence".strip(),
        ]
        evidence_block = await _evidence_block(queries, context, max_facts=18)
        data = await llm_json(
            system=_risk_generation_system_prompt(),
            user_msg=(
                f"{_project_context_text(context)}\n\n"
                f"Confirmed categories:\n{category_text}"
                f"{evidence_block}\n\n"
                f"Produce {MIN_RISKS_PER_CATEGORY}-{MAX_RISKS_PER_CATEGORY} risks for every category above.\n"
                "Return JSON only with key risks. Each risk object must include exactly these keys: "
                "title, category, affected_components, why_it_matters, evidence_basis, "
                "missing_information, evidence_status."
            ),
            model=settings.openai_orchestration_model,
            context=context,
        )
        risks = _normalize_risks(data.get("risks") or [], category_labels)
        risks = await self._fill_thin_categories(
            context, category_labels, category_text, evidence_block, risks
        )
        return _cap_per_category(risks, MAX_RISKS_PER_CATEGORY)

    async def _fill_thin_categories(
        self,
        context: dict,
        category_labels: list[str],
        category_text: str,
        evidence_block: str,
        risks: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        """Ask the model once more for categories that came back under-covered.

        Depth is topped up dynamically rather than from static templates, so
        unusual category labels (market, hydrology, supply chain, …) are handled
        the same as conventional development-finance ones.
        """
        thin = {
            label: MIN_RISKS_PER_CATEGORY - _count_for_category(risks, label)
            for label in category_labels
            if _count_for_category(risks, label) < MIN_RISKS_PER_CATEGORY
        }
        if not thin:
            return risks

        wanted = "\n".join(f"- {label}: {needed} more" for label, needed in thin.items())
        existing_titles = "\n".join(f"- {risk['title']}" for risk in risks) or "- (none yet)"
        try:
            data = await llm_json(
                system=_risk_generation_system_prompt(),
                user_msg=(
                    f"{_project_context_text(context)}\n\n"
                    f"Categories:\n{category_text}"
                    f"{evidence_block}\n\n"
                    f"Risks already drafted (do not repeat or rephrase these):\n{existing_titles}\n\n"
                    f"These categories still need more risks:\n{wanted}\n\n"
                    "Return JSON only with key risks, using the same keys as before, "
                    "covering only the categories listed above."
                ),
                model=settings.openai_orchestration_model,
                context=context,
            )
        except Exception:
            logger.exception("Risk depth top-up failed; keeping first-pass risks")
            return risks

        additions = _normalize_risks(data.get("risks") or [], list(thin.keys()))
        seen = {_dedupe_key(risk["title"]) for risk in risks}
        for addition in additions:
            label = addition["category"]
            if thin.get(label, 0) <= 0:
                continue
            key = _dedupe_key(addition["title"])
            if key in seen:
                continue
            seen.add(key)
            risks.append(addition)
            thin[label] -= 1
        return risks

    async def _enrich_mitigation(
        self,
        item_content: dict,
        existing_record: dict,
        context: dict,
    ) -> dict:
        queries = [
            f"mitigation {item_content.get('title', '')} {context.get('project_type', '')} {context.get('geography', '')}".strip()
        ]
        evidence_block = await _evidence_block(queries, context, max_facts=8)
        data = await llm_json(
            system=(
                "You are a practical project risk mitigation specialist. Propose mitigation for the "
                "specific risk. Avoid vague statements like 'monitor closely'. Return JSON with "
                "keys mitigation, owner, timing, remaining_issue, status. Status must be one of "
                "Adequate, Insufficient, Needs validation."
            ),
            user_msg=(
                f"{_project_context_text(context)}\n\n"
                f"Risk: {item_content.get('title', '')}\n"
                f"Category: {item_content.get('category', '')}\n"
                f"Why it matters: {item_content.get('why_it_matters', item_content.get('description', ''))}\n"
                f"Evidence/basis: {item_content.get('evidence_basis', item_content.get('basis', ''))}\n"
                f"Missing information: {item_content.get('missing_information', '')}"
                f"{evidence_block}"
            ),
            model=settings.openai_orchestration_model,
            context=context,
        )
        return {
            "mitigation": data.get("mitigation", existing_record.get("mitigation", "")),
            "owner": data.get("owner", existing_record.get("owner", "")),
            "timing": data.get("timing", existing_record.get("timing", "")),
            "remaining_issue": data.get("remaining_issue", existing_record.get("remaining_issue", "")),
            "status": data.get("status", existing_record.get("status", "Needs validation")),
        }

    async def _enrich_mitigations_bulk(
        self,
        source_items: list[dict[str, Any]],
        existing_records: dict[str, dict[str, Any]],
        context: dict,
    ) -> dict[str, dict[str, Any]]:
        risk_payload = [
            {
                "source_item_id": item.get("id", ""),
                "title": item.get("content", {}).get("title", ""),
                "category": item.get("content", {}).get("category", ""),
                "affected_components": item.get("content", {}).get("affected_components", ""),
                "why_it_matters": item.get("content", {}).get("why_it_matters", item.get("content", {}).get("description", "")),
                "evidence_basis": item.get("content", {}).get("evidence_basis", item.get("content", {}).get("basis", "")),
                "missing_information": item.get("content", {}).get("missing_information", ""),
            }
            for item in source_items
        ]
        descriptor = _project_descriptor(context)
        queries = [
            f"{descriptor} risk mitigation project preparation implementation controls".strip(),
            f"{descriptor} risk mitigation procurement institutional data verification safeguards".strip(),
        ]
        evidence_block = await _evidence_block(queries, context, max_facts=12)
        data = await llm_json(
            system=(
                "You are a practical development-finance project-preparation specialist. "
                "For each risk, propose a specific mitigation that addresses the stated cause and consequence. "
                "Avoid vague statements like 'monitor closely' or 'strengthen coordination' unless you specify "
                "the concrete mechanism, owner, and timing. Do not invent unsupported institutional facts. "
                "Return JSON only with key mitigations: objects with source_item_id, mitigation, owner, timing, "
                "remaining_issue, status. Status must be Adequate, Insufficient, or Needs validation."
            ),
            user_msg=(
                f"{_project_context_text(context)}\n\n"
                f"Risks needing mitigation:\n{_jsonish(risk_payload)}"
                f"{evidence_block}"
            ),
            model=settings.openai_orchestration_model,
            context=context,
        )
        by_id = {
            str(item.get("source_item_id", "")): item
            for item in data.get("mitigations", [])
            if isinstance(item, dict)
        }

        records: dict[str, dict[str, Any]] = {}
        for item in source_items:
            item_id = str(item.get("id", ""))
            existing = existing_records.get(item_id, {})
            llm_record = by_id.get(item_id, {})
            content = item.get("content", {})
            fallback = _default_mitigation(content, context)
            records[item_id] = {
                "mitigation": llm_record.get("mitigation") or existing.get("mitigation") or fallback["mitigation"],
                "owner": llm_record.get("owner") or existing.get("owner") or fallback["owner"],
                "timing": llm_record.get("timing") or existing.get("timing") or fallback["timing"],
                "remaining_issue": llm_record.get("remaining_issue") or existing.get("remaining_issue") or fallback["remaining_issue"],
                "status": llm_record.get("status") or existing.get("status") or fallback["status"],
            }
        return records

    async def _rate_register(
        self,
        context: dict,
        category_items: list[dict],
        draft_register: list[dict[str, Any]],
    ) -> dict[str, Any]:
        data = await llm_json(
            system=(
                "You are a senior development-finance risk reviewer. Assign qualitative ratings only "
                "now that risks and mitigations are known. Use exactly one of Low, Moderate, "
                "Substantial, High. Ratings are structured judgment, not mathematical certainty. "
                "Return JSON with keys category_ratings, risk_register, top_risks, unresolved_issues. "
                "For each risk_register row include risk_id, inherent_rating, residual_rating, "
                "rating_rationale, basis_evidence, missing_information."
            ),
            user_msg=(
                f"{_project_context_text(context)}\n\n"
                f"Categories:\n{_jsonish([c.get('content', {}) for c in category_items])}\n\n"
                f"Draft register:\n{_jsonish(draft_register)}"
            ),
            model=settings.openai_orchestration_model,
            context=context,
        )
        return data or {}

    def _build_register_seed(
        self,
        risk_items: list[dict],
        mitigation_records: dict[str, dict],
    ) -> list[dict[str, Any]]:
        rows = []
        for idx, item in enumerate(risk_items, start=1):
            content = item.get("content", {})
            record = mitigation_records.get(item.get("id", ""), {}) or {}
            rows.append({
                "risk_id": f"R{idx:02d}",
                "source_item_id": item.get("id", ""),
                "category": content.get("category", ""),
                "risk_title": content.get("title", content.get("risk", "")),
                "description": content.get("why_it_matters", content.get("description", "")),
                "affected_components": content.get("affected_components", ""),
                "inherent_rating": "Moderate",
                "mitigation": record.get("mitigation", ""),
                "residual_rating": "Moderate",
                "owner_status": _owner_status(record),
                "basis_evidence": content.get("evidence_basis", content.get("basis", "")),
                "missing_information": content.get("missing_information", ""),
                "rating_rationale": "",
                "remaining_issue": record.get("remaining_issue", ""),
                "mitigation_status": record.get("status", ""),
            })
        return rows

    def _derive_category_ratings(
        self,
        category_items: list[dict],
        register_rows: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        ratings = []
        for item in category_items:
            label = item.get("content", {}).get("label", "")
            rows = [row for row in register_rows if row.get("category") == label]
            if not label or not rows:
                continue
            rating = max((row.get("residual_rating", "Moderate") for row in rows), key=_rating_score)
            ratings.append({
                "category": label,
                "rating": rating,
                "rationale": "Derived from the highest residual rating among confirmed risks in this category.",
                "top_risks": [row.get("risk_title", "") for row in rows[:3]],
                "unresolved_issues": [
                    row.get("missing_information", "")
                    for row in rows
                    if row.get("missing_information")
                ],
            })
        return ratings

    def _derive_top_risks(self, register_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        sorted_rows = sorted(
            register_rows,
            key=lambda row: (_rating_score(row.get("residual_rating")), _rating_score(row.get("inherent_rating"))),
            reverse=True,
        )
        return [
            {
                "risk_id": row.get("risk_id", ""),
                "risk_title": row.get("risk_title", ""),
                "why_it_matters": row.get("rating_rationale") or row.get("basis_evidence", ""),
                "mitigation_summary": row.get("mitigation", ""),
            }
            for row in sorted_rows[:5]
        ]

    def _derive_unresolved_issues(self, register_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [
            {
                "risk_id": row.get("risk_id", ""),
                "issue": row.get("missing_information") or row.get("remaining_issue", ""),
            }
            for row in register_rows
            if row.get("missing_information") or row.get("remaining_issue")
        ]

    @staticmethod
    def export_xlsx(widget_data: dict[str, Any]) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for XLSX export") from exc

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risk Register"

        headers = [
            "Risk ID",
            "Category",
            "Risk Title",
            "Description",
            "Affected Components",
            "Inherent Rating",
            "Mitigation",
            "Residual Rating",
            "Owner / Status",
            "Basis / Evidence",
            "Missing Information",
            "Rating Rationale",
        ]
        _write_sheet(ws, headers, [
            [
                row.get("risk_id", ""),
                row.get("category", ""),
                row.get("risk_title", ""),
                row.get("description", ""),
                row.get("affected_components", ""),
                row.get("inherent_rating", ""),
                row.get("mitigation", ""),
                row.get("residual_rating", ""),
                row.get("owner_status", ""),
                row.get("basis_evidence", ""),
                row.get("missing_information", ""),
                row.get("rating_rationale", ""),
            ]
            for row in widget_data.get("risk_register", [])
        ], Font, PatternFill, Alignment)

        category_ws = wb.create_sheet("Category Ratings")
        _write_sheet(category_ws, ["Category", "Rating", "Rationale", "Top Risks", "Unresolved Issues"], [
            [
                row.get("category", ""),
                row.get("rating", ""),
                row.get("rationale", ""),
                "\n".join(_as_list(row.get("top_risks"))),
                "\n".join(_as_list(row.get("unresolved_issues"))),
            ]
            for row in widget_data.get("category_ratings", [])
        ], Font, PatternFill, Alignment)

        top_ws = wb.create_sheet("Top Risks")
        _write_sheet(top_ws, ["Risk ID", "Risk", "Why It Matters", "Mitigation Summary"], [
            [
                row.get("risk_id", ""),
                row.get("risk_title", ""),
                row.get("why_it_matters", ""),
                row.get("mitigation_summary", ""),
            ]
            for row in widget_data.get("top_risks", [])
        ], Font, PatternFill, Alignment)

        issues_ws = wb.create_sheet("Unresolved Issues")
        _write_sheet(issues_ws, ["Risk ID", "Issue"], [
            [row.get("risk_id", ""), row.get("issue", "")]
            for row in widget_data.get("unresolved_issues", [])
        ], Font, PatternFill, Alignment)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


def _write_sheet(ws, headers, rows, Font, PatternFill, Alignment) -> None:
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="E6F4F1")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 42)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


async def _evidence_block(queries: list[str], context: dict[str, Any], max_facts: int = 15) -> str:
    db = context.get("_db")
    project_id = context.get("project_id")
    if db is None or not project_id:
        return ""
    context_str, _citations = await retrieve_evidence(
        [q for q in queries if q],
        db,
        project_id,
        max_facts=max_facts,
        user_id=context.get("user_id"),
    )
    if not context_str:
        return ""
    return f"\n\nRetrieved evidence and context. Use only if relevant; cite/support in basis fields:\n{context_str}"


def _context_queries(context: dict, suffix: str) -> list[str]:
    return [
        f"{context.get('project_type', '')} {context.get('geography', '')} {suffix}".strip(),
        f"{context.get('project_title', '')} {suffix}".strip(),
    ]


def _project_context_text(context: dict) -> str:
    return (
        f"Project title: {context.get('project_title', '')}\n"
        f"Project type/sector: {context.get('project_type', '')}\n"
        f"Geography: {context.get('geography', '')}\n"
        f"Target population: {context.get('target_population', '')}\n"
        f"Description: {context.get('project_description', '')}\n"
        f"Project plan: {_jsonish(context.get('project_plan', {}))}\n"
        f"Prior tool/assessment inputs: {_jsonish(context.get('tool_inputs', {}))}"
    )


def _project_descriptor(context: dict[str, Any]) -> str:
    """Short retrieval descriptor built only from this project's own context.

    Kept free of sector keyword taxonomies so retrieval works the same for any
    geography, sector, or delivery model.
    """
    parts = [
        str(context.get("project_title", "")).strip(),
        str(context.get("project_type", "")).replace("_", " ").strip(),
        str(context.get("geography", "")).strip(),
        str(context.get("target_population", "")).strip(),
    ]
    return " ".join(part for part in parts if part)


def _risk_generation_system_prompt() -> str:
    return (
        "You are a senior development-finance risk reviewer preparing a project-specific risk assessment. "
        "Generate fewer, better risks: 2-4 high-quality risks per confirmed category.\n\n"
        "Risk title requirements:\n"
        "- Write concrete risk statements, not labels.\n"
        "- Use a cause -> consequence pattern: '[specific condition] could [specific consequence] for [project component/outcome].'\n"
        "- Avoid labels like 'Data Fragmentation Risk' or 'Complex stakeholder engagement process'.\n"
        "- Prefer falsifiable statements a reviewer can accept, reject, or mark as needing evidence.\n\n"
        "Grounding requirements:\n"
        "- Use project context aggressively: country, sector, delivery model, beneficiaries, technologies, agencies, documents, and precedents when available.\n"
        "- Do not invent unsupported facts. If a fact is not evidenced, phrase it as uncertainty, e.g. 'Unclear licensing requirements could delay contracting or approvals.'\n"
        "- Separate the broad category from the project-specific manifestation in the fields.\n"
        "- Evidence/basis must say what project material, retrieved source, or explicit variable supports the risk.\n\n"
        "Internal specificity check before returning JSON: reject or rewrite any risk if it could apply to almost any project, "
        "lacks a clear consequence, does not connect to project context, implies unsupported facts, or duplicates another risk.\n\n"
        "Return JSON only: {\"risks\": [{\"title\", \"category\", \"affected_components\", "
        "\"why_it_matters\", \"evidence_basis\", \"missing_information\", \"evidence_status\"}]}."
    )


def _default_categories(context: dict) -> list[dict[str, str]]:
    project_type = context.get("project_type", "project")
    return [
        {"label": "Political and Governance", "description": "Government decision-making and policy continuity.", "why_it_matters": "Public approvals and support can affect implementation.", "status": "Include"},
        {"label": "Sector Policy and Regulatory", "description": "Rules, tariffs, permits, and sector oversight.", "why_it_matters": f"{project_type} projects often depend on enabling sector rules.", "status": "Include"},
        {"label": "Technical Design and Delivery", "description": "Technology, design, construction, and operational delivery risks.", "why_it_matters": "Delivery complexity can affect cost, schedule, and performance.", "status": "Include"},
        {"label": "Institutional Capacity", "description": "Implementing agency and partner capability.", "why_it_matters": "Capacity constraints can slow procurement, coordination, and execution.", "status": "Include"},
        {"label": "Fiduciary and Procurement", "description": "Financial management, procurement, and contracting risks.", "why_it_matters": "Weak controls can delay or compromise implementation.", "status": "Include"},
        {"label": "Environmental and Social", "description": "Community, land, environmental, labor, and safeguard risks.", "why_it_matters": "E&S issues can create harm, opposition, or compliance gaps.", "status": "Include"},
        {"label": "Data Quality and Results Verification", "description": "Evidence, monitoring, MRV, and verification risks.", "why_it_matters": "Weak data can undermine results claims and donor confidence.", "status": "Include"},
    ]


def _default_mitigation(risk: dict[str, Any], context: dict[str, Any]) -> dict[str, str]:
    title = str(risk.get("title", "")).lower()
    category = str(risk.get("category", "")).lower()
    affected = str(risk.get("affected_components", "affected workstream")).strip() or "affected workstream"
    project_team = "Project team / implementing agency"
    if "data" in title or "verification" in title or "mrv" in category:
        return {
            "mitigation": (
                "Define a single data dictionary, source-of-truth owner, and QA process before using "
                f"{affected} data for targeting, implementation planning, or results reporting."
            ),
            "owner": project_team,
            "timing": "Preparation / before baseline lock",
            "remaining_issue": "Underlying data availability and cross-agency access rights still need confirmation.",
            "status": "Needs validation",
        }
    if "permit" in title or "approval" in title or "policy" in category or "regulatory" in category:
        return {
            "mitigation": (
                "Map required approvals, assign each approval to a named owner, and build approval lead times "
                "into the procurement and rollout schedule."
            ),
            "owner": project_team,
            "timing": "Preparation / before procurement launch",
            "remaining_issue": "Current approval requirements and agency decision timelines need evidence.",
            "status": "Needs validation",
        }
    if "procurement" in title or "contract" in title or "fiduciary" in category:
        return {
            "mitigation": (
                "Prepare a procurement plan with bid-package sequencing, evaluation criteria, contract KPIs, "
                "and escalation steps for delayed awards."
            ),
            "owner": "Procurement lead / implementing agency",
            "timing": "Preparation through contract award",
            "remaining_issue": "Supplier market depth and procurement authority capacity need validation.",
            "status": "Needs validation",
        }
    if "capacity" in title or "institutional" in category:
        return {
            "mitigation": (
                "Assign delivery roles across agencies and partners, define decision rights, and resource a PMO "
                "or coordination unit to track milestones and unblock dependencies."
            ),
            "owner": project_team,
            "timing": "Preparation / early implementation",
            "remaining_issue": "Partner staffing, budget, and mandate clarity need confirmation.",
            "status": "Needs validation",
        }
    return {
        "mitigation": (
            f"Create an owner-assigned action plan for {affected}, with evidence needed, decision owner, "
            "target date, and escalation trigger before the risk can affect implementation."
        ),
        "owner": project_team,
        "timing": "Preparation",
        "remaining_issue": "Specific validation evidence and accountable owner need confirmation.",
        "status": "Needs validation",
    }


def _normalize_risks(
    raw_risks: list[Any],
    category_labels: list[str],
) -> list[dict[str, Any]]:
    """Normalize model output into risk rows, dropping placeholders and duplicates.

    Screening is deliberately structural (is this a real, substantive risk
    statement?) rather than keyword-based: keyword gates rejected valid risks for
    any project whose wording did not echo its own sector/geography terms.
    """
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for raw in raw_risks:
        if not isinstance(raw, dict):
            continue
        title = str(raw.get("title") or raw.get("risk") or "").strip()
        why = str(raw.get("why_it_matters") or raw.get("description") or "").strip()
        if not _has_risk_substance(title, why):
            continue
        key = _dedupe_key(title)
        if key in seen:
            continue
        seen.add(key)
        normalized.append({
            "title": title,
            "category": _category_or_default(raw.get("category"), category_labels),
            "affected_components": _join_if_list(raw.get("affected_components", "")),
            "why_it_matters": why,
            "evidence_basis": str(
                raw.get("evidence_basis") or raw.get("basis") or raw.get("rationale") or ""
            ).strip(),
            "missing_information": _join_if_list(raw.get("missing_information", "")),
            "evidence_status": _normalize_evidence_status(raw.get("evidence_status")),
        })
    return normalized


def _has_risk_substance(title: str, why_it_matters: str) -> bool:
    """A usable risk needs a real statement and a stated consequence."""
    if _is_placeholder_risk_title(title):
        return False
    # Bare category labels ("Procurement risk") are too short to be falsifiable.
    if len(title) < 20 or len(title.split()) < 4:
        return False
    return bool(why_it_matters)


def _count_for_category(risks: list[dict[str, Any]], label: str) -> int:
    return sum(1 for risk in risks if risk.get("category") == label)


def _cap_per_category(risks: list[dict[str, Any]], maximum: int) -> list[dict[str, Any]]:
    counts: dict[str, int] = {}
    capped: list[dict[str, Any]] = []
    for risk in risks:
        label = str(risk.get("category", ""))
        if counts.get(label, 0) >= maximum:
            continue
        counts[label] = counts.get(label, 0) + 1
        capped.append(risk)
    return capped


def _dedupe_key(title: str) -> str:
    return " ".join(str(title or "").lower().split())


def _normalize_evidence_status(value: Any) -> str:
    candidate = str(value or "").strip().lower()
    for status in EVIDENCE_STATUSES:
        if candidate == status.lower():
            return status
    return "Partially supported"


def _is_placeholder_risk_title(value: Any) -> bool:
    title = str(value or "").strip().lower()
    if not title:
        return True
    placeholder_phrases = (
        "variables need validation",
        "variable needs validation",
        "risk to be confirmed",
        "to be confirmed",
        "tbd",
        "needs validation",
        "category-specific execution risk",
        "unresolved variables",
    )
    return any(phrase in title for phrase in placeholder_phrases)


def _find_llm_row(seed: dict[str, Any], rows: list[dict[str, Any]], idx: int) -> dict[str, Any] | None:
    risk_id = seed.get("risk_id")
    title = str(seed.get("risk_title", "")).strip().lower()
    for row in rows:
        if row.get("risk_id") == risk_id:
            return row
        if title and str(row.get("risk_title", row.get("title", ""))).strip().lower() == title:
            return row
    return rows[idx] if idx < len(rows) and isinstance(rows[idx], dict) else None


def _normalize_rating(value: Any) -> str:
    normalized = str(value or "").strip().lower()
    for rating in RISK_RATINGS:
        if normalized == rating.lower():
            return rating
    if normalized in {"medium", "moderate risk"}:
        return "Moderate"
    if normalized in {"significant", "substantial risk"}:
        return "Substantial"
    return "Moderate"


def _rating_score(value: Any) -> int:
    try:
        return RISK_RATINGS.index(_normalize_rating(value))
    except ValueError:
        return 1


def _owner_status(record: dict[str, Any]) -> str:
    owner = str(record.get("owner", "")).strip()
    status = str(record.get("status", "")).strip()
    if owner and status:
        return f"{owner} / {status}"
    return owner or status


def _category_or_default(value: Any, valid_categories: list[str]) -> str:
    """Snap a model-supplied category onto a confirmed label.

    Falls back to the first confirmed category so a mislabelled row still lands
    somewhere the reviewer can see and re-file it.
    """
    category = str(value or "").strip()
    if category in valid_categories:
        return category
    lowered = category.lower()
    for label in valid_categories:
        if label.lower() == lowered:
            return label
    return valid_categories[0] if valid_categories else category


def _as_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(v) for v in value if str(v).strip()]
    if value is None:
        return []
    text = str(value).strip()
    return [text] if text else []


def _join_if_list(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value if str(v).strip())
    return str(value or "")


def _jsonish(value: Any) -> str:
    import json
    return json.dumps(value, ensure_ascii=True, default=str, indent=2)


def _risk_register_markdown(rows: list[dict[str, Any]]) -> str:
    headers = ["Risk ID", "Category", "Risk", "Inherent", "Mitigation", "Residual", "Owner / Status", "Missing Info"]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(_clean_cell(row.get(key, "")) for key in [
            "risk_id",
            "category",
            "risk_title",
            "inherent_rating",
            "mitigation",
            "residual_rating",
            "owner_status",
            "missing_information",
        ]) + " |")
    return "\n".join(lines)


def _risk_register_tsv(rows: list[dict[str, Any]]) -> str:
    keys = [
        "risk_id",
        "category",
        "risk_title",
        "description",
        "affected_components",
        "inherent_rating",
        "mitigation",
        "residual_rating",
        "owner_status",
        "basis_evidence",
        "missing_information",
    ]
    lines = ["\t".join(keys)]
    for row in rows:
        lines.append("\t".join(str(row.get(key, "")).replace("\t", " ").replace("\n", " ") for key in keys))
    return "\n".join(lines)


def _clean_cell(value: Any) -> str:
    return str(value or "").replace("|", "\\|").replace("\n", " ").strip()
