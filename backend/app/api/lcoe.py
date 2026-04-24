"""LCOE API endpoints — recalculate, sensitivity, and Excel export."""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Any
import io
import logging

from app.core.auth import get_current_user, MockUser
from app.modules.lcoe_module import LCOETool
from app.services.lcoe_engine import LCOEEngine, LCOEInput

router = APIRouter()
logger = logging.getLogger(__name__)


class RecalculateRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]


class SensitivityRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]
    params: list[str] | None = None
    delta: float = 0.20


class UpdateInputRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]
    field_name: str
    value: Any
    source: str = "user"
    status: str = "validated"


def _normalize_input_status(status: str) -> str:
    return "validated" if status == "confirmed" else status


@router.post("/lcoe/recalculate")
async def recalculate_lcoe(
    data: RecalculateRequest,
    user: MockUser = Depends(get_current_user),
):
    """Recalculate LCOE from a full input set. No LLM call — pure math."""
    tool = LCOETool()
    try:
        result = await tool.recalculate(data.inputs)
        return result
    except (ValueError, ZeroDivisionError):
        raise HTTPException(status_code=400, detail="Invalid inputs for LCOE calculation")


@router.post("/lcoe/update-input")
async def update_input_and_recalculate(
    data: UpdateInputRequest,
    user: MockUser = Depends(get_current_user),
):
    """Update a single input field and recalculate."""
    inputs = data.inputs
    normalized_status = _normalize_input_status(data.status)
    if data.field_name in inputs:
        inputs[data.field_name]["value"] = data.value
        inputs[data.field_name]["source"] = data.source
        inputs[data.field_name]["status"] = normalized_status
    else:
        inputs[data.field_name] = {
            "field_name": data.field_name,
            "label": data.field_name,
            "value": data.value,
            "unit": "",
            "source": data.source,
            "status": normalized_status,
            "notes": "",
            "rationale": "",
            "category": "general",
        }

    tool = LCOETool()
    try:
        result = await tool.recalculate(inputs)
        return result
    except (ValueError, ZeroDivisionError):
        raise HTTPException(status_code=400, detail="Invalid inputs for LCOE calculation")


@router.post("/lcoe/sensitivity")
async def run_sensitivity(
    data: SensitivityRequest,
    user: MockUser = Depends(get_current_user),
):
    """Run sensitivity analysis on the LCOE model."""
    engine_inputs = {k: LCOEInput.from_dict(v) for k, v in data.inputs.items()}
    try:
        points = LCOEEngine.run_sensitivity(
            engine_inputs,
            params=data.params,
            delta=data.delta,
        )
        return {"sensitivity": [p.to_dict() for p in points]}
    except (ValueError, ZeroDivisionError):
        raise HTTPException(status_code=400, detail="Invalid inputs for sensitivity analysis")


@router.post("/lcoe/export")
async def export_lcoe_excel(
    data: RecalculateRequest,
    user: MockUser = Depends(get_current_user),
):
    """Export the LCOE model as a formula-driven Excel workbook.

    Sheets: Model (main), Cash Flows, Inputs, Sensitivity.
    Key inputs live on the Model sheet — change them and the entire
    workbook recalculates.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    engine_inputs = {k: LCOEInput.from_dict(v) for k, v in data.inputs.items()}

    if not LCOEEngine.is_computable(engine_inputs):
        raise HTTPException(status_code=400, detail="Not enough inputs to compute LCOE")

    try:
        result = LCOEEngine.calculate(engine_inputs)
    except (ValueError, ZeroDivisionError):
        raise HTTPException(status_code=400, detail="Invalid inputs for LCOE calculation")

    sensitivity_points: list = []
    try:
        sensitivity_points = LCOEEngine.run_sensitivity(engine_inputs)
    except Exception:
        pass

    wb = openpyxl.Workbook()

    # -- Shared styles --
    BLUE = "004D91"
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    section_font = Font(bold=True, size=11, color=BLUE)
    title_font = Font(bold=True, size=14, color=BLUE)
    input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    assumed_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    base_highlight = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    double_top = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="double"), bottom=Side(style="thin"),
    )
    pct_fmt = "0.0%"
    cur_fmt = "#,##0.00"
    int_fmt = "#,##0"

    def _val(name: str, fallback: float = 0.0) -> float:
        inp = data.inputs.get(name, {})
        v = inp.get("value")
        if v is None:
            return fallback
        try:
            return float(v)
        except (ValueError, TypeError):
            return fallback

    capacity_kw = _val("net_capacity_kw")
    capacity_factor = _val("capacity_factor")
    total_capex = _val("total_capex")
    annual_opex = _val("annual_opex")
    annual_fuel = _val("annual_fuel_cost", 0.0)
    annual_replacement = _val("annual_replacement_cost", 0.0)
    discount_rate = _val("discount_rate")
    project_life = int(_val("project_life_years", 25))
    construction_years = int(_val("construction_years", 0))
    degradation = _val("degradation_rate", 0.005)
    currency = "USD"
    if "currency" in data.inputs:
        currency = str(data.inputs["currency"].get("value", "USD") or "USD")
    total_years = construction_years + project_life

    # ================================================================
    # SHEET 1 — Model  (main tab with editable inputs + formula results)
    # ================================================================
    ws = wb.active
    ws.title = "Model"
    ws.sheet_properties.tabColor = BLUE

    ws.merge_cells("A1:C1")
    ws["A1"].value = "LCOE Financial Model"
    ws["A1"].font = title_font

    # Fixed row layout for stable cross-sheet references
    R_SEC_INPUTS = 3
    R_CAP = 4
    R_CF = 5
    R_CAPEX = 6
    R_OPEX = 7
    R_FUEL = 8
    R_REPL = 9
    R_DR = 10
    R_LIFE = 11
    R_CONSTR = 12
    R_DEG = 13

    ws.cell(R_SEC_INPUTS, 1, "KEY INPUTS").font = section_font

    input_defs = [
        (R_CAP,    "Net Capacity",             capacity_kw,         "kW",             None),
        (R_CF,     "Capacity Factor",           capacity_factor,     "",               pct_fmt),
        (R_CAPEX,  "Total CAPEX",               total_capex,         currency,         cur_fmt),
        (R_OPEX,   "Annual O&M",                annual_opex,         f"{currency}/yr", cur_fmt),
        (R_FUEL,   "Annual Fuel Cost",           annual_fuel,         f"{currency}/yr", cur_fmt),
        (R_REPL,   "Annual Replacement Cost",    annual_replacement,  f"{currency}/yr", cur_fmt),
        (R_DR,     "Discount Rate (WACC)",       discount_rate,       "",               pct_fmt),
        (R_LIFE,   "Project Lifetime",           project_life,        "years",          int_fmt),
        (R_CONSTR, "Construction Period",         construction_years,  "years",          int_fmt),
        (R_DEG,    "Degradation Rate",           degradation,         "%/yr",           pct_fmt),
    ]
    for row, label, value, unit, nfmt in input_defs:
        ws.cell(row, 1, label).font = Font(size=11)
        c = ws.cell(row, 2, value)
        c.fill = input_fill
        c.border = thin_border
        if nfmt:
            c.number_format = nfmt
        if unit:
            ws.cell(row, 3, unit).font = Font(color="808080", size=9)

    # Derived values
    R_SEC_DERIVED = 15
    R_ENERGY = 16
    R_CAPEX_YR = 17
    ws.cell(R_SEC_DERIVED, 1, "DERIVED VALUES").font = section_font

    ws.cell(R_ENERGY, 1, "Base Annual Energy")
    ws.cell(R_ENERGY, 2).value = f"=B{R_CAP}*B{R_CF}*8760"
    ws.cell(R_ENERGY, 2).number_format = int_fmt
    ws.cell(R_ENERGY, 3, "kWh/yr").font = Font(color="808080", size=9)

    ws.cell(R_CAPEX_YR, 1, "CAPEX per Construction Year")
    ws.cell(R_CAPEX_YR, 2).value = f"=B{R_CAPEX}/MAX(B{R_CONSTR},1)"
    ws.cell(R_CAPEX_YR, 2).number_format = cur_fmt
    ws.cell(R_CAPEX_YR, 3, currency).font = Font(color="808080", size=9)

    # Results (formulas wired after Cash Flows sheet is built)
    R_SEC_RESULTS = 19
    R_NPV_COST = 20
    R_NPV_ENERGY = 21
    R_LCOE = 22
    ws.cell(R_SEC_RESULTS, 1, "RESULTS").font = section_font

    ws.cell(R_NPV_COST, 1, "NPV Total Costs")
    ws.cell(R_NPV_COST, 2).number_format = cur_fmt
    ws.cell(R_NPV_COST, 3, currency).font = Font(color="808080", size=9)
    ws.cell(R_NPV_ENERGY, 1, "NPV Total Energy")
    ws.cell(R_NPV_ENERGY, 2).number_format = int_fmt
    ws.cell(R_NPV_ENERGY, 3, "kWh").font = Font(color="808080", size=9)
    ws.cell(R_LCOE, 1, "LCOE")
    ws.cell(R_LCOE, 2).value = f"=B{R_NPV_COST}/B{R_NPV_ENERGY}"
    ws.cell(R_LCOE, 2).font = Font(bold=True, size=13, color=BLUE)
    ws.cell(R_LCOE, 2).number_format = "0.000000"
    ws.cell(R_LCOE, 3, f"{currency}/kWh").font = Font(color="808080", size=9)

    # Cost breakdown (formulas wired after Cash Flows sheet is built)
    R_SEC_BREAK = 24
    R_CAPEX_SH = 25
    R_OPEX_SH = 26
    R_FUEL_SH = 27
    R_REPL_SH = 28
    ws.cell(R_SEC_BREAK, 1, "COST BREAKDOWN (NPV)").font = section_font
    ws.cell(R_CAPEX_SH, 1, "CAPEX Share")
    ws.cell(R_OPEX_SH, 1, "O&M Share")
    ws.cell(R_FUEL_SH, 1, "Fuel Share")
    ws.cell(R_REPL_SH, 1, "Replacement Share")
    for r in (R_CAPEX_SH, R_OPEX_SH, R_FUEL_SH, R_REPL_SH):
        ws.cell(r, 2).number_format = pct_fmt

    # Quality
    R_QUALITY = 30
    ws.cell(R_QUALITY, 1, "Assumptions Used")
    ws.cell(R_QUALITY, 2, result.assumption_count)
    ws.cell(R_QUALITY + 1, 1, "Estimate Quality")
    ws.cell(R_QUALITY + 1, 2, result.quality_label).font = Font(
        bold=True,
        color="1B5E20" if result.quality_label == "high"
        else "F57F17" if result.quality_label == "moderate"
        else "B71C1C",
    )

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14

    # ================================================================
    # SHEET 2 — Cash Flows  (fully formula-driven)
    # ================================================================
    ws_cf = wb.create_sheet("Cash Flows")

    cf_headers = [
        "Year", "CAPEX", "O&M", "Fuel", "Replacement",
        "Total Cost", "Energy (kWh)", "Discount Factor",
        "Discounted Cost", "Discounted Energy",
    ]
    for col, h in enumerate(cf_headers, 1):
        c = ws_cf.cell(1, col, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border

    CF_START = 2
    for i in range(total_years):
        r = CF_START + i
        ws_cf.cell(r, 1, i).border = thin_border                                        # A: Year

        ws_cf.cell(r, 2).value = f"=IF(A{r}<Model!$B${R_CONSTR},Model!$B${R_CAPEX_YR},0)"  # B: CAPEX
        ws_cf.cell(r, 2).border = thin_border
        ws_cf.cell(r, 2).number_format = cur_fmt

        ws_cf.cell(r, 3).value = f"=IF(A{r}>=Model!$B${R_CONSTR},Model!$B${R_OPEX},0)"    # C: O&M
        ws_cf.cell(r, 3).border = thin_border
        ws_cf.cell(r, 3).number_format = cur_fmt

        ws_cf.cell(r, 4).value = f"=IF(A{r}>=Model!$B${R_CONSTR},Model!$B${R_FUEL},0)"    # D: Fuel
        ws_cf.cell(r, 4).border = thin_border
        ws_cf.cell(r, 4).number_format = cur_fmt

        ws_cf.cell(r, 5).value = f"=IF(A{r}>=Model!$B${R_CONSTR},Model!$B${R_REPL},0)"    # E: Replacement
        ws_cf.cell(r, 5).border = thin_border
        ws_cf.cell(r, 5).number_format = cur_fmt

        ws_cf.cell(r, 6).value = f"=B{r}+C{r}+D{r}+E{r}"                                  # F: Total Cost
        ws_cf.cell(r, 6).border = thin_border
        ws_cf.cell(r, 6).number_format = cur_fmt

        energy_formula = (                                                                   # G: Energy
            f"=IF(A{r}>=Model!$B${R_CONSTR},"
            f"Model!$B${R_ENERGY}*(1-Model!$B${R_DEG})^(A{r}-Model!$B${R_CONSTR}),0)"
        )
        ws_cf.cell(r, 7).value = energy_formula
        ws_cf.cell(r, 7).border = thin_border
        ws_cf.cell(r, 7).number_format = cur_fmt

        df_formula = f"=IF(Model!$B${R_DR}>0,1/(1+Model!$B${R_DR})^A{r},1)"                # H: Discount Factor
        ws_cf.cell(r, 8).value = df_formula
        ws_cf.cell(r, 8).border = thin_border
        ws_cf.cell(r, 8).number_format = "0.000000"

        ws_cf.cell(r, 9).value = f"=F{r}*H{r}"                                             # I: Disc. Cost
        ws_cf.cell(r, 9).border = thin_border
        ws_cf.cell(r, 9).number_format = cur_fmt

        ws_cf.cell(r, 10).value = f"=G{r}*H{r}"                                            # J: Disc. Energy
        ws_cf.cell(r, 10).border = thin_border
        ws_cf.cell(r, 10).number_format = cur_fmt

    CF_END = CF_START + total_years - 1

    # Totals row
    TOT_ROW = CF_END + 2
    ws_cf.cell(TOT_ROW, 1, "TOTAL").font = Font(bold=True)
    for col_idx in range(2, 11):
        col_letter = chr(64 + col_idx)
        c = ws_cf.cell(TOT_ROW, col_idx)
        c.value = f"=SUM({col_letter}{CF_START}:{col_letter}{CF_END})"
        c.font = Font(bold=True)
        c.border = double_top
        c.number_format = cur_fmt

    # LCOE computed from cash-flow sums
    LCOE_CF_ROW = TOT_ROW + 1
    ws_cf.cell(LCOE_CF_ROW, 1, "LCOE").font = Font(bold=True, size=11, color=BLUE)
    ws_cf.cell(LCOE_CF_ROW, 2).value = f"=I{TOT_ROW}/J{TOT_ROW}"
    ws_cf.cell(LCOE_CF_ROW, 2).font = Font(bold=True, size=11, color=BLUE)
    ws_cf.cell(LCOE_CF_ROW, 2).number_format = "0.000000"
    ws_cf.cell(LCOE_CF_ROW, 3, f"{currency}/kWh").font = Font(color="808080", size=9)

    # Discounted component sums (used by Model sheet cost-breakdown)
    DISC_HDR = LCOE_CF_ROW + 2
    ws_cf.cell(DISC_HDR, 1, "Discounted Component Totals").font = Font(bold=True, italic=True)
    DISC_CAPEX_R = DISC_HDR + 1
    DISC_OPEX_R = DISC_HDR + 2
    DISC_FUEL_R = DISC_HDR + 3
    DISC_REPL_R = DISC_HDR + 4
    for row_off, label, src_col in [
        (DISC_CAPEX_R, "Discounted CAPEX", "B"),
        (DISC_OPEX_R,  "Discounted O&M",   "C"),
        (DISC_FUEL_R,  "Discounted Fuel",   "D"),
        (DISC_REPL_R,  "Discounted Replacement", "E"),
    ]:
        ws_cf.cell(row_off, 1, label)
        ws_cf.cell(row_off, 2).value = (
            f"=SUMPRODUCT({src_col}{CF_START}:{src_col}{CF_END},"
            f"H{CF_START}:H{CF_END})"
        )
        ws_cf.cell(row_off, 2).number_format = cur_fmt

    for col in range(1, 11):
        ws_cf.column_dimensions[chr(64 + col)].width = 18

    # -- Wire Model sheet formulas that reference Cash Flows --
    ws.cell(R_NPV_COST, 2).value = f"='Cash Flows'!I{TOT_ROW}"
    ws.cell(R_NPV_ENERGY, 2).value = f"='Cash Flows'!J{TOT_ROW}"

    ws.cell(R_CAPEX_SH, 2).value = f"=IF(B{R_NPV_COST}>0,'Cash Flows'!B{DISC_CAPEX_R}/B{R_NPV_COST},0)"
    ws.cell(R_OPEX_SH, 2).value = f"=IF(B{R_NPV_COST}>0,'Cash Flows'!B{DISC_OPEX_R}/B{R_NPV_COST},0)"
    ws.cell(R_FUEL_SH, 2).value = f"=IF(B{R_NPV_COST}>0,'Cash Flows'!B{DISC_FUEL_R}/B{R_NPV_COST},0)"
    ws.cell(R_REPL_SH, 2).value = f"=IF(B{R_NPV_COST}>0,'Cash Flows'!B{DISC_REPL_R}/B{R_NPV_COST},0)"

    # ================================================================
    # SHEET 3 — Inputs  (full provenance table)
    # ================================================================
    ws_inp = wb.create_sheet("Inputs")
    inp_headers = ["Field", "Value", "Unit", "Source", "Status", "Notes"]
    for col, h in enumerate(inp_headers, 1):
        c = ws_inp.cell(1, col, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border

    for row_idx, inp_data in enumerate(data.inputs.values(), 2):
        vals = [
            inp_data.get("label", inp_data.get("field_name", "")),
            inp_data.get("value", ""),
            inp_data.get("unit", ""),
            inp_data.get("source", ""),
            inp_data.get("status", ""),
            inp_data.get("notes", ""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws_inp.cell(row_idx, col, v)
            c.border = thin_border
            if inp_data.get("status") == "assumed":
                c.fill = assumed_fill

    for col in range(1, 7):
        ws_inp.column_dimensions[chr(64 + col)].width = 22

    # ================================================================
    # SHEET 4 — Sensitivity  (server-computed)
    # ================================================================
    if sensitivity_points:
        ws_sens = wb.create_sheet("Sensitivity")

        sens_by_param: dict[str, list] = {}
        for pt in sensitivity_points:
            sens_by_param.setdefault(pt.param_name, []).append(pt)

        r = 1
        for _param_name, points in sens_by_param.items():
            points.sort(key=lambda p: p.test_value)
            ws_sens.cell(r, 1, points[0].param_label).font = Font(bold=True, size=11)
            r += 1
            for col, h in enumerate(["Parameter Value", f"LCOE ({currency}/kWh)", "Δ from Base"], 1):
                c = ws_sens.cell(r, col, h)
                c.font = header_font_white
                c.fill = header_fill
                c.border = thin_border
            r += 1
            base_lcoe = result.lcoe
            for pt in points:
                ws_sens.cell(r, 1, pt.test_value).border = thin_border
                ws_sens.cell(r, 2, pt.lcoe).border = thin_border
                ws_sens.cell(r, 2).number_format = "0.000000"
                delta = ((pt.lcoe - base_lcoe) / base_lcoe) if base_lcoe else 0
                ws_sens.cell(r, 3, delta).border = thin_border
                ws_sens.cell(r, 3).number_format = "+0.0%;-0.0%;0.0%"
                is_base = abs(pt.test_value - pt.base_value) < abs(pt.base_value) * 0.001
                if is_base:
                    for c_idx in range(1, 4):
                        ws_sens.cell(r, c_idx).fill = base_highlight
                r += 1
            r += 1

        ws_sens.column_dimensions["A"].width = 20
        ws_sens.column_dimensions["B"].width = 24
        ws_sens.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=lcoe_model.xlsx"},
    )
