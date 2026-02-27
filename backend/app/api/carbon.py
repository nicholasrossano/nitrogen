"""Carbon Emissions Calculator API — recalculate, sensitivity, and Excel export."""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Any
import io
import logging

from app.core.database import get_db
from app.core.auth import get_current_user, MockUser
from app.tools.carbon_tool import CarbonTool
from app.services.carbon_engine import CarbonEngine, CarbonInput

router = APIRouter()
logger = logging.getLogger(__name__)


class RecalculateRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]


class SensitivityRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]
    params: list[str] | None = None
    delta: float = 0.20


class UpdateInputRequest(BaseModel):
    inputs: dict[str, dict[str, Any]]
    field_name: str
    value: Any
    source: str = "user"
    status: str = "confirmed"


@router.post("/carbon/recalculate")
async def recalculate_carbon(
    data: RecalculateRequest,
    user: MockUser = Depends(get_current_user),
):
    """Recalculate carbon ERs from a full input set. No LLM call — pure math."""
    tool = CarbonTool()
    try:
        result = await tool.recalculate(data.inputs)
        return result
    except (ValueError, ZeroDivisionError) as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/carbon/update-input")
async def update_input_and_recalculate(
    data: UpdateInputRequest,
    user: MockUser = Depends(get_current_user),
):
    """Update a single input field and recalculate."""
    inputs = data.inputs
    if data.field_name in inputs:
        inputs[data.field_name]["value"] = data.value
        inputs[data.field_name]["source"] = data.source
        inputs[data.field_name]["status"] = data.status
    else:
        inputs[data.field_name] = {
            "field_name": data.field_name,
            "label": data.field_name,
            "value": data.value,
            "unit": "",
            "source": data.source,
            "status": data.status,
            "applies_to": "general",
            "notes": "",
            "rationale": "",
            "category": "general",
        }

    tool = CarbonTool()
    try:
        result = await tool.recalculate(inputs)
        return result
    except (ValueError, ZeroDivisionError) as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/carbon/sensitivity")
async def run_sensitivity(
    data: SensitivityRequest,
    user: MockUser = Depends(get_current_user),
):
    """Run sensitivity analysis on the carbon model."""
    engine_inputs = {k: CarbonInput.from_dict(v) for k, v in data.inputs.items()}
    try:
        points = CarbonEngine.run_sensitivity(
            engine_inputs,
            params=data.params,
            delta=data.delta,
        )
        return {"sensitivity": [p.to_dict() for p in points]}
    except (ValueError, ZeroDivisionError) as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/carbon/export")
async def export_carbon_excel(
    data: RecalculateRequest,
    user: MockUser = Depends(get_current_user),
):
    """Export the carbon ER model as a formula-driven Excel workbook.

    Sheets: Model (main), ER Schedule, Inputs, Sensitivity.
    Key inputs live on the Model sheet — change them and the entire
    workbook recalculates.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    engine_inputs = {k: CarbonInput.from_dict(v) for k, v in data.inputs.items()}

    if not CarbonEngine.is_computable(engine_inputs):
        raise HTTPException(status_code=400, detail="Not enough inputs to compute emission reductions")

    try:
        result = CarbonEngine.calculate(engine_inputs)
    except (ValueError, ZeroDivisionError) as e:
        raise HTTPException(status_code=400, detail=str(e))

    sensitivity_points: list = []
    try:
        sensitivity_points = CarbonEngine.run_sensitivity(engine_inputs)
    except Exception:
        pass

    wb = openpyxl.Workbook()

    # -- Shared styles --
    GREEN = "1B5E20"
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    section_font = Font(bold=True, size=11, color=GREEN)
    title_font = Font(bold=True, size=14, color=GREEN)
    input_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    assumed_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    base_highlight = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    double_top = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="double"), bottom=Side(style="thin"),
    )
    dec4_fmt = "0.0000"
    pct_fmt = "0.0%"
    int_fmt = "#,##0"

    def _val(name: str, fallback: float = 0.0) -> float:
        inp = data.inputs.get(name, {})
        v = inp.get("value")
        if v is None:
            return fallback
        try:
            return float(v)
        except (ValueError, TypeError):
            return fallback

    def _str(name: str, fallback: str = "") -> str:
        inp = data.inputs.get(name, {})
        v = inp.get("value")
        return str(v) if v is not None else fallback

    devices = _val("devices_households")
    usage_rate = _val("usage_rate", 1.0)
    adoption_rate = _val("adoption_rate", 1.0)
    bl_fuel_kg = _val("baseline_fuel_consumption_kg_yr")
    pj_fuel_kg = _val("project_fuel_consumption_kg_yr")
    bl_ncv = _val("baseline_ncv_mj_kg", 15.6)
    pj_ncv = _val("project_ncv_mj_kg", 15.6)
    bl_eff = _val("baseline_efficiency", 0.10)
    pj_eff = _val("project_efficiency", 0.30)
    ef_tco2_per_tj = _val("emission_factor_tco2_per_tj", 112.0)
    ef_kgco2_per_kg = _val("emission_factor_kgco2_per_kg", 0.0)
    fnrb = _val("fnrb", 0.70)
    leakage_factor = _val("leakage_factor", 0.0)
    fuel_savings_pct = _val("fuel_savings_pct", 0.0)
    crediting_years = int(_val("crediting_period_years", 10))
    project_is_biomass = _str("project_is_biomass", "true").lower() in ("true", "yes", "1")

    # Derive effective project fuel (mirrors engine logic)
    effective_pj_fuel = pj_fuel_kg
    pj_fuel_note = ""
    if effective_pj_fuel == 0 and bl_fuel_kg > 0:
        if fuel_savings_pct > 0:
            effective_pj_fuel = bl_fuel_kg * (1 - fuel_savings_pct)
            pj_fuel_note = f"Derived: baseline × (1 − {fuel_savings_pct:.0%} savings)"
        elif bl_eff > 0 and pj_eff > 0:
            effective_pj_fuel = bl_fuel_kg * (bl_eff / pj_eff)
            pj_fuel_note = f"Derived: baseline × (eff {bl_eff:.0%} / {pj_eff:.0%})"

    # Per-device emission factors (pre-computed — branching logic)
    if ef_kgco2_per_kg > 0:
        bl_tco2_per_device = bl_fuel_kg * ef_kgco2_per_kg / 1000
        pj_tco2_per_device = effective_pj_fuel * ef_kgco2_per_kg / 1000
        ef_note = f"Direct: fuel_kg × {ef_kgco2_per_kg} kgCO₂/kg ÷ 1000"
    else:
        bl_tco2_per_device = (bl_fuel_kg * bl_ncv / 1_000_000) * ef_tco2_per_tj
        pj_tco2_per_device = (effective_pj_fuel * pj_ncv / 1_000_000) * ef_tco2_per_tj
        ef_note = f"NCV chain: fuel_kg × NCV ÷ 1e6 × {ef_tco2_per_tj} tCO₂/TJ"

    fnrb_project = fnrb if project_is_biomass else 1.0

    # ================================================================
    # SHEET 1 — Model  (main tab)
    # ================================================================
    ws = wb.active
    ws.title = "Model"
    ws.sheet_properties.tabColor = GREEN

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Carbon Emission Reductions Model"
    ws["A1"].font = title_font

    # Fixed row layout
    R_SEC_INPUTS = 3
    R_DEV = 4; R_USG = 5; R_ADPT = 6
    R_BL_FUEL = 7; R_PJ_FUEL = 8
    R_FNRB = 9; R_FNRB_PJ = 10
    R_LEAK = 11; R_CREDIT = 12

    ws.cell(R_SEC_INPUTS, 1, "KEY INPUTS").font = section_font

    input_defs = [
        (R_DEV,     "Devices / Households",       devices,           "units",             int_fmt),
        (R_USG,     "Usage Rate",                  usage_rate,        "",                  pct_fmt),
        (R_ADPT,    "Adoption Rate",               adoption_rate,     "",                  pct_fmt),
        (R_BL_FUEL, "Baseline Fuel Consumption",   bl_fuel_kg,        "kg/yr per device",  "#,##0.0"),
        (R_PJ_FUEL, "Project Fuel Consumption",    effective_pj_fuel, "kg/yr per device",  "#,##0.0"),
        (R_FNRB,    "fNRB",                        fnrb,              "",                  pct_fmt),
        (R_FNRB_PJ, "fNRB (project side)",         fnrb_project,      "",                  pct_fmt),
        (R_LEAK,    "Leakage Factor",              leakage_factor,    "",                  pct_fmt),
        (R_CREDIT,  "Crediting Period",            crediting_years,   "years",             int_fmt),
    ]
    for row, label, value, unit, nfmt in input_defs:
        ws.cell(row, 1, label).font = Font(size=11)
        c = ws.cell(row, 2, value)
        c.fill = input_fill
        c.border = thin_border
        if nfmt:
            c.number_format = nfmt
        if unit:
            ws.cell(row, 3, unit).font = Font(color="808080", size=9)

    if pj_fuel_note:
        ws.cell(R_PJ_FUEL, 4, pj_fuel_note).font = Font(color="808080", italic=True, size=9)
    fnrb_pj_note = "= fNRB (project uses biomass)" if project_is_biomass else "= 1.0 (non-biomass project)"
    ws.cell(R_FNRB_PJ, 4, fnrb_pj_note).font = Font(color="808080", italic=True, size=9)

    # Emission factors (pre-computed)
    R_SEC_EF = 14
    R_BL_EF = 15; R_PJ_EF = 16
    ws.cell(R_SEC_EF, 1, "EMISSION FACTORS (per device)").font = section_font

    ws.cell(R_BL_EF, 1, "Baseline (tCO₂e/device/yr)")
    ws.cell(R_BL_EF, 2, bl_tco2_per_device).number_format = "0.000000"
    ws.cell(R_BL_EF, 2).border = thin_border
    ws.cell(R_BL_EF, 3, ef_note).font = Font(color="808080", italic=True, size=9)

    ws.cell(R_PJ_EF, 1, "Project (tCO₂e/device/yr)")
    ws.cell(R_PJ_EF, 2, pj_tco2_per_device).number_format = "0.000000"
    ws.cell(R_PJ_EF, 2).border = thin_border

    # Year 1 results (formulas)
    R_SEC_YR1 = 18
    R_BL_EM = 19; R_PJ_EM = 20; R_LK_EM = 21; R_NET = 22
    ws.cell(R_SEC_YR1, 1, "YEAR 1 RESULTS").font = section_font

    ws.cell(R_BL_EM, 1, "Baseline Emissions (tCO₂e)")
    ws.cell(R_BL_EM, 2).value = f"=B{R_DEV}*B{R_USG}*B{R_BL_EF}*B{R_FNRB}"
    ws.cell(R_BL_EM, 2).number_format = dec4_fmt

    ws.cell(R_PJ_EM, 1, "Project Emissions (tCO₂e)")
    ws.cell(R_PJ_EM, 2).value = f"=B{R_DEV}*B{R_USG}*B{R_PJ_EF}*B{R_FNRB_PJ}"
    ws.cell(R_PJ_EM, 2).number_format = dec4_fmt

    ws.cell(R_LK_EM, 1, "Leakage (tCO₂e)")
    ws.cell(R_LK_EM, 2).value = f"=B{R_LEAK}*MAX(B{R_BL_EM}-B{R_PJ_EM},0)"
    ws.cell(R_LK_EM, 2).number_format = dec4_fmt

    ws.cell(R_NET, 1, "Net Emission Reductions (tCO₂e)")
    ws.cell(R_NET, 2).value = f"=B{R_BL_EM}-B{R_PJ_EM}-B{R_LK_EM}"
    ws.cell(R_NET, 2).font = Font(bold=True, size=13, color=GREEN)
    ws.cell(R_NET, 2).number_format = dec4_fmt
    ws.cell(R_NET, 3, "tCO₂e/yr").font = Font(color="808080", size=9)

    # Crediting period totals (formulas wired after ER Schedule is built)
    R_SEC_TOT = 24
    R_TOT_BL = 25; R_TOT_PJ = 26; R_TOT_LK = 27; R_TOT_NET = 28
    ws.cell(R_SEC_TOT, 1, "CREDITING PERIOD TOTALS").font = section_font
    ws.cell(R_TOT_BL, 1, "Total Baseline Emissions")
    ws.cell(R_TOT_BL, 2).number_format = dec4_fmt
    ws.cell(R_TOT_PJ, 1, "Total Project Emissions")
    ws.cell(R_TOT_PJ, 2).number_format = dec4_fmt
    ws.cell(R_TOT_LK, 1, "Total Leakage")
    ws.cell(R_TOT_LK, 2).number_format = dec4_fmt
    ws.cell(R_TOT_NET, 1, "Total Net ERs")
    ws.cell(R_TOT_NET, 2).font = Font(bold=True, size=12, color=GREEN)
    ws.cell(R_TOT_NET, 2).number_format = dec4_fmt
    ws.cell(R_TOT_NET, 3, "tCO₂e").font = Font(color="808080", size=9)

    # Quality
    R_QUALITY = 30
    ws.cell(R_QUALITY, 1, "Assumptions Used")
    ws.cell(R_QUALITY, 2, result.assumption_count)
    ws.cell(R_QUALITY + 1, 1, "Estimate Quality")
    ws.cell(R_QUALITY + 1, 2, result.quality_label).font = Font(
        bold=True,
        color="1B5E20" if result.quality_label == "high"
        else "F57F17" if result.quality_label == "moderate"
        else "B71C1C",
    )

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40

    # ================================================================
    # SHEET 2 — ER Schedule  (fully formula-driven)
    # ================================================================
    ws_er = wb.create_sheet("ER Schedule")

    er_headers = [
        "Year", "Devices Active", "Baseline Emissions (tCO₂e)",
        "Project Emissions (tCO₂e)", "Leakage (tCO₂e)", "Net ERs (tCO₂e)",
    ]
    for col, h in enumerate(er_headers, 1):
        c = ws_er.cell(1, col, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border

    ER_START = 2
    for yr in range(1, crediting_years + 1):
        r = ER_START + yr - 1

        ws_er.cell(r, 1, yr).border = thin_border  # A: Year

        # B: Devices Active — mirrors engine adoption logic
        devices_formula = (
            f"=INT(Model!$B${R_DEV}*IF(Model!$B${R_ADPT}<1,"
            f"MIN(Model!$B${R_ADPT}*A{r},1),Model!$B${R_ADPT}))"
        )
        ws_er.cell(r, 2).value = devices_formula
        ws_er.cell(r, 2).border = thin_border
        ws_er.cell(r, 2).number_format = int_fmt

        # C: Baseline = devices × usage × EF_bl × fNRB
        ws_er.cell(r, 3).value = (
            f"=B{r}*Model!$B${R_USG}*Model!$B${R_BL_EF}*Model!$B${R_FNRB}"
        )
        ws_er.cell(r, 3).border = thin_border
        ws_er.cell(r, 3).number_format = dec4_fmt

        # D: Project = devices × usage × EF_pj × fNRB_pj
        ws_er.cell(r, 4).value = (
            f"=B{r}*Model!$B${R_USG}*Model!$B${R_PJ_EF}*Model!$B${R_FNRB_PJ}"
        )
        ws_er.cell(r, 4).border = thin_border
        ws_er.cell(r, 4).number_format = dec4_fmt

        # E: Leakage = leakage_factor × MAX(baseline − project, 0)
        ws_er.cell(r, 5).value = f"=Model!$B${R_LEAK}*MAX(C{r}-D{r},0)"
        ws_er.cell(r, 5).border = thin_border
        ws_er.cell(r, 5).number_format = dec4_fmt

        # F: Net ERs = baseline − project − leakage
        ws_er.cell(r, 6).value = f"=C{r}-D{r}-E{r}"
        ws_er.cell(r, 6).border = thin_border
        ws_er.cell(r, 6).number_format = dec4_fmt

    ER_END = ER_START + crediting_years - 1

    # Totals row
    TOT_ROW = ER_END + 2
    ws_er.cell(TOT_ROW, 1, "TOTAL").font = Font(bold=True)
    for col_idx in range(2, 7):
        col_letter = chr(64 + col_idx)
        c = ws_er.cell(TOT_ROW, col_idx)
        c.value = f"=SUM({col_letter}{ER_START}:{col_letter}{ER_END})"
        c.font = Font(bold=True)
        c.border = double_top
        c.number_format = dec4_fmt

    for col in range(1, 7):
        ws_er.column_dimensions[chr(64 + col)].width = 24

    # -- Wire Model sheet totals to ER Schedule sums --
    ws.cell(R_TOT_BL, 2).value = f"='ER Schedule'!C{TOT_ROW}"
    ws.cell(R_TOT_PJ, 2).value = f"='ER Schedule'!D{TOT_ROW}"
    ws.cell(R_TOT_LK, 2).value = f"='ER Schedule'!E{TOT_ROW}"
    ws.cell(R_TOT_NET, 2).value = f"='ER Schedule'!F{TOT_ROW}"

    # ================================================================
    # SHEET 3 — Inputs  (full provenance table)
    # ================================================================
    ws_inp = wb.create_sheet("Inputs")
    inp_headers = ["Field", "Value", "Unit", "Applies To", "Source", "Status", "Notes"]
    for col, h in enumerate(inp_headers, 1):
        c = ws_inp.cell(1, col, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border

    for row_idx, inp_data in enumerate(data.inputs.values(), 2):
        vals = [
            inp_data.get("label", inp_data.get("field_name", "")),
            inp_data.get("value", ""),
            inp_data.get("unit", ""),
            inp_data.get("applies_to", ""),
            inp_data.get("source", ""),
            inp_data.get("status", ""),
            inp_data.get("notes", ""),
        ]
        for col, v in enumerate(vals, 1):
            c = ws_inp.cell(row_idx, col, v)
            c.border = thin_border
            if inp_data.get("status") == "assumed":
                c.fill = assumed_fill

    for col in range(1, 8):
        ws_inp.column_dimensions[chr(64 + col)].width = 22

    # ================================================================
    # SHEET 4 — Sensitivity  (server-computed)
    # ================================================================
    if sensitivity_points:
        ws_sens = wb.create_sheet("Sensitivity")

        sens_by_param: dict[str, list] = {}
        for pt in sensitivity_points:
            sens_by_param.setdefault(pt.param_name, []).append(pt)

        r = 1
        for _param_name, points in sens_by_param.items():
            points.sort(key=lambda p: p.test_value)
            ws_sens.cell(r, 1, points[0].param_label).font = Font(bold=True, size=11)
            r += 1
            for col, h in enumerate(["Parameter Value", "Net ERs (tCO₂e)", "Δ from Base"], 1):
                c = ws_sens.cell(r, col, h)
                c.font = header_font_white
                c.fill = header_fill
                c.border = thin_border
            r += 1
            base_net_er = result.net_er_tco2e
            for pt in points:
                ws_sens.cell(r, 1, pt.test_value).border = thin_border
                ws_sens.cell(r, 2, pt.net_er).border = thin_border
                ws_sens.cell(r, 2).number_format = dec4_fmt
                delta = ((pt.net_er - base_net_er) / base_net_er) if base_net_er else 0
                ws_sens.cell(r, 3, delta).border = thin_border
                ws_sens.cell(r, 3).number_format = "+0.0%;-0.0%;0.0%"
                is_base = abs(pt.test_value - pt.base_value) < abs(pt.base_value) * 0.001
                if is_base:
                    for c_idx in range(1, 4):
                        ws_sens.cell(r, c_idx).fill = base_highlight
                r += 1
            r += 1

        ws_sens.column_dimensions["A"].width = 20
        ws_sens.column_dimensions["B"].width = 24
        ws_sens.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=carbon_er_model.xlsx"},
    )
