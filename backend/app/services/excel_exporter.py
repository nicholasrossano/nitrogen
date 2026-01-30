"""Excel export service for checklists."""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporterService:
    """Service for exporting checklists to Excel."""
    
    def __init__(self):
        self.exports_dir = "/tmp/exports"
        os.makedirs(self.exports_dir, exist_ok=True)
        
        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="7677B8", end_color="7677B8", fill_type="solid")
        self.category_font = Font(bold=True, size=11)
        self.category_fill = PatternFill(start_color="ECE4DB", end_color="ECE4DB", fill_type="solid")
        
        self.risk_fills = {
            "low": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
            "medium": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            "high": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        }
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    async def export_checklist(self, content: dict) -> str:
        """Export checklist to Excel file."""
        wb = Workbook()
        
        # Main checklist sheet
        ws = wb.active
        ws.title = "Checklist"
        
        # Set column widths
        ws.column_dimensions['A'].width = 8   # Status
        ws.column_dimensions['B'].width = 45  # Item
        ws.column_dimensions['C'].width = 12  # Risk Level
        ws.column_dimensions['D'].width = 50  # Questions
        ws.column_dimensions['E'].width = 30  # Notes
        
        # Title row
        ws.merge_cells('A1:E1')
        ws['A1'] = content.get('title', 'Due Diligence Checklist')
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date and risk rating
        ws.merge_cells('A2:E2')
        risk_rating = content.get('overall_risk_rating', 'medium')
        ws['A2'] = f"Generated: {content.get('date', datetime.now().strftime('%Y-%m-%d'))} | Overall Risk: {risk_rating.upper()}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        row = 4
        headers = ['Status', 'Checklist Item', 'Risk', 'Key Questions', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        row += 1
        
        # Categories and items
        categories = content.get('categories', [])
        for category in categories:
            # Category header
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=category.get('name', ''))
            cell.font = self.category_font
            cell.fill = self.category_fill
            cell.border = self.thin_border
            row += 1
            
            # Items
            for item in category.get('items', []):
                # Status column (checkbox placeholder)
                cell = ws.cell(row=row, column=1, value='☐')
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.thin_border
                
                # Item description
                cell = ws.cell(row=row, column=2, value=item.get('item', ''))
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.thin_border
                
                # Risk level
                risk = item.get('risk_level', 'medium')
                cell = ws.cell(row=row, column=3, value=risk.upper())
                cell.alignment = Alignment(horizontal='center')
                cell.fill = self.risk_fills.get(risk, self.risk_fills['medium'])
                cell.border = self.thin_border
                
                # Questions
                questions = item.get('questions', [])
                cell = ws.cell(row=row, column=4, value='\n'.join(f"• {q}" for q in questions))
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.thin_border
                
                # Notes (empty for user to fill)
                cell = ws.cell(row=row, column=5, value='')
                cell.border = self.thin_border
                
                # Adjust row height based on content
                ws.row_dimensions[row].height = max(30, 15 * (len(questions) + 1))
                
                row += 1
            
            row += 1  # Space between categories
        
        # Priority Items sheet
        if content.get('priority_items'):
            ws_priority = wb.create_sheet("Priority Items")
            ws_priority.column_dimensions['A'].width = 8
            ws_priority.column_dimensions['B'].width = 60
            ws_priority.column_dimensions['C'].width = 30
            
            ws_priority['A1'] = "Priority Items to Address First"
            ws_priority['A1'].font = Font(bold=True, size=14)
            ws_priority.merge_cells('A1:C1')
            
            row = 3
            headers = ['#', 'Priority Item', 'Status/Notes']
            for col, header in enumerate(headers, 1):
                cell = ws_priority.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
            
            row += 1
            for idx, item in enumerate(content['priority_items'], 1):
                ws_priority.cell(row=row, column=1, value=idx).border = self.thin_border
                ws_priority.cell(row=row, column=2, value=item).border = self.thin_border
                ws_priority.cell(row=row, column=3, value='').border = self.thin_border
                row += 1
        
        # Next Steps sheet
        if content.get('next_steps'):
            ws_steps = wb.create_sheet("Next Steps")
            ws_steps.column_dimensions['A'].width = 8
            ws_steps.column_dimensions['B'].width = 60
            ws_steps.column_dimensions['C'].width = 20
            ws_steps.column_dimensions['D'].width = 30
            
            ws_steps['A1'] = "Recommended Next Steps"
            ws_steps['A1'].font = Font(bold=True, size=14)
            ws_steps.merge_cells('A1:D1')
            
            row = 3
            headers = ['#', 'Action Item', 'Due Date', 'Owner']
            for col, header in enumerate(headers, 1):
                cell = ws_steps.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
            
            row += 1
            for idx, step in enumerate(content['next_steps'], 1):
                ws_steps.cell(row=row, column=1, value=idx).border = self.thin_border
                ws_steps.cell(row=row, column=2, value=step).border = self.thin_border
                ws_steps.cell(row=row, column=3, value='').border = self.thin_border
                ws_steps.cell(row=row, column=4, value='').border = self.thin_border
                row += 1
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"due_diligence_checklist_{timestamp}.xlsx"
        filepath = os.path.join(self.exports_dir, filename)
        wb.save(filepath)
        
        return filepath
